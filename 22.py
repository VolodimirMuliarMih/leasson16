#xlxs file
import csv
with open('tel_csv') as f:
    full_list = list(csv.reader(f))
import openpyxl
wb = openpyxl.Workbook()
wb.create_sheet(title = "Первая страница", index=0)
wb.remove(wb['Sheet'])
sheet_work = wb["Первая страница"]
for cel_index, cell in enumerate ((full_list)):
    if cel_index==2:
        pass
    for row_index, value in enumerate (cell):
        cell = sheet_work.cell(column=cel_index+1, row = row_index+1)
        cell.value = value
wb.save('task_17.xlsx')
