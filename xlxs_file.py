#xlxs file
import csv
with open('tel_csv') as f:
    full_list = list(csv.reader(f))
    list_titlf = full_list[0]
    list_petr = full_list[2]
    list_irina = full_list[4]
    list_oleg = full_list[6]
    list_genediy = full_list[8]
    del list_titlf[2]
    del list_petr[2]
    del list_irina[2]
    del list_oleg [2]
    del list_genediy[2]
import openpyxl
wb = openpyxl.Workbook()
wb.create_sheet(title = "Первая страница", index=0)
wb.remove(wb['Sheet'])
sheet_work = wb["Первая страница"]
for cel_index, cell in enumerate ((list_titlf,list_petr,list_irina,list_oleg,list_genediy)):
    for row_index, value in enumerate (cell):
        cell = sheet_work.cell(column=cel_index+1, row = row_index+1)
        cell.value = value
wb.save('task_17.xlsx')

